import openpyxl

wb = openpyxl.load_workbook(r'c:\NotebookLM\Lista_Cotejo_Supervision_EMS (PW SEP2025).xlsx')

for sheet_name in wb.sheetnames[:3]:
    ws = wb[sheet_name]
    print(f'=== Sheet: {sheet_name} ===')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    for i, row in enumerate(ws.iter_rows(max_row=min(25, ws.max_row))):
        cells = []
        for c in row:
            val = str(c.value)[:50] if c.value is not None else ''
            cells.append(val)
        print(f'  Row {i+1}: {cells}')
    print()
